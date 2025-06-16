import os
import logging
import re
from collections import defaultdict
from datetime import datetime
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# ✅ 中文 EPU 關鍵詞
E_TERMS = ["GDP", "ISM", "PCE", "PPI", "不動產", "不動產市場", "不動產業", "中國內需市場", "中國股票", "太陽光電", "太陽能", "半導體", "半導體產業", "台股", "台商", "台積電", "外送運價", "外匯", "外匯存底", "外資", "外銷", "石化", "光電推動", "全球貿易體系", "地熱", "住宅價格指數", "汽車產業", "併購", "供應鏈", "房市", "房地產", "房貸", "房價", "服務業", "股市", "股匯", "金控集團", "金融", "金融市場", "金融機構", "政經情勢", "科技領域", "美元", "重工業", "個人消費支出物價指數", "旅宿業", "旅遊業", "核電", "海洋能", "海運", "消費力道", "消費市場", "消費性電子市場", "能源", "財測", "高科技業", "高耗能產業", "商業", "國內景氣", "國內經濟景氣", "國防業務", "國家競爭力", "國銀人民幣存款餘額", "國際貿易", "淨零碳排", "淨零轉型", "產業轉移", "勞動力人口", "就業市場", "就業指數", "就業報告數據", "景氣", "晶片產業", "晶片業", "無人機產業", "貿易", "傳產", "傳統產業", "匯市", "經貿", "經貿活動", "經貿發展", "經濟", "經濟成長", "經濟委員會", "經濟體", "電力", "電價", "碳費", "綠色科技新創產業", "綠能", "綠電", "製造業", "製造業指數", "數位轉型", "機械產業", "鋼市", "鋼鐵業", "營建", "營建業"]

P_TERMS = ["CBAM", "CPI", "CPTPP", "Fed", "FED", "IMF", "WTO", "人行", "中央銀行", "中國人民銀行", "中國加Ｎ", "中國加一", "中國商務部", "中國國有銀行", "中國國務院", "中經院", "升息", "升準", "友岸外包", "友岸合作", "反間諜法", "反傾銷稅", "反補貼稅", "日本銀行", "日銀", "世銀", "出口擴大", "加徵", "加徵稅", "加徵關稅", "外國投資", "央行", "打炒房", "打炒房措施", "申報碳排放係數", "全球政策", "存量房貸利率", "低價傾銷", "利率正常化", "利率會議", "刺激措施", "房貸利率", "房貸緊縮", "拋匯", "非農就業", "保護主義", "信用管制", "信用管制措施", "政策打炒房", "政策改革", "政策寬鬆", "政策調控", "美國降息", "重貼現率", "降息", "降準", "風電政策", "修正草案", "振興經濟", "振興經濟措施", "核能政策", "消費者物價指數", "特殊主權債券", "租稅抵減", "缺工會議", "財政收支", "財政刺激", "財政政策", "財政與貨幣不確定性", "財政寬鬆", "國務院", "國產化", "國際貨幣基金", "國際貿易規則", "基準利率", "掠奪性定價", "採取措施", "救市", "救市政策", "救市措施", "淨零", "淨零政策", "淨零碳排", "深度節能", "產業電價調漲", "移工政策", "第七波打炒房", "貨幣互換協定", "貨幣和信用政策", "貨幣政策", "貨幣與財政刺激", "貨幣寬鬆", "通貨緊縮", "創造青年就業", "創造就業", "就業津貼補助", "提高企業稅", "最低工資", "殖利率", "減半調漲", "發行國債", "稅務優惠", "貿易管制", "開徵碳費", "傾銷", "新三樣", "新青安", "補貼", "補貼措施", "資金緊縮", "跨太平洋夥伴全面進步協定", "零關稅", "零關稅優惠", "電費", "電費調漲", "電價拍板", "電價拍板定案", "電價政策", "電價費率", "電價調漲", "電價調整", "預算", "實施凍漲", "監管政策", "碳定價", "碳費", "碳費制度", "碳費定價", "碳費費率審議", "碳價優惠", "碳邊境調整機制", "緊縮", "銀行存款準備率", "銀行限貸", "增加國債", "寬鬆貨幣", "徵收碳費", "撥補", "歐盟執委會", "課稅", "課徵金融所得稅", "課徵關稅", "調漲最低薪資", "選擇性信用管制", "優惠政策", "優惠費率", "總量管制", "總預算", "聯準會", "聯準會降息", "懲罰性關稅", "關稅優惠", "川普2.0", "貿易政策", "對等關稅", "關稅措施", "關稅戰", "川普關稅", "房貸限縮", "關稅政策", "關稅貿易戰", "關稅提高"]

U_TERMS = ["人口高齡化", "人心惶惶", "大失所望", "大幅衰退", "干擾", "不友善勢力", "不可投資", "不平靜", "不如預期", "不安", "不明", "不振", "不景氣", "不會太好", "不確定", "不確定因素", "不確定性", "不確定風險", "不適合投資", "不穩定", "中共封控威脅", "中東地緣政治", "中東局勢", "中東緊張局勢", "中東戰火", "中東變局", "內需低迷", "內需消費低迷", "反效果", "反彈", "反覆無常", "少子女化", "少子化", "少子化危機", "引發外界質疑", "引進移工", "充滿變數", "加速逃離", "失去信心", "失望", "失業", "失業率居高", "失業率居高不下", "失業率高", "失業率飆高", "左右為難", "市況慘淡", "市場不安", "市場失望", "市場風險", "市場震盪", "打擊消費者信心", "民眾抗爭", "由升轉貶", "示警", "全球貿易壁壘", "全球衝突", "全球競爭加劇", "再審議", "地緣政治", "地緣政治崛起", "地緣政治緊張", "成本上揚", "成本挑戰", "成長放緩", "成長率放緩", "有待觀察", "考驗", "低於預期", "低迷不振", "冷戰", "妨礙", "局勢不安", "局勢緊張", "局勢緊繃", "技術人才短缺", "投資人疑慮", "抗議", "攻防", "材料供不應求", "沉重負擔", "供不應求", "供應鏈風險", "受阻", "尚未有進一步的安排", "尚待觀察", "房市正崩盤", "房市泡沫", "房市信心大受衝擊", "房市崩跌", "房市崩盤與經濟前景不確定", "房地產危機持續延燒", "房貸不降反增", "房貸緊縮", "房價持續崩跌", "放緩", "油價下滑", "治安不佳", "波動", "金龍海嘯", "阻撓", "俄烏戰爭", "信心低迷", "信心低靡", "信心受挫", "削弱熱情", "前景不確定", "威脅升溫", "威脅性", "客戶需求的變化", "封控", "待改進", "很難經營", "急凍", "持續惡化", "政治干擾", "政治考量備受關注", "政治風險", "政治紛擾", "政治動盪", "政策不確定性", "政策反覆無常", "洩密疑慮", "看壞", "美中科技紛爭", "美中貿易戰", "美東罷工", "美國大選恐將翻盤", "美國大選逼近", "美國總統大選", "苦戰", "訂單大幅減少", "負面效應", "負面影響", "重挫", "風險上升", "倒閉", "恐怕", "恐將翻盤", "效果難料", "消費力道降低", "消費力道將會降低", "消費低迷", "消極", "疲軟", "缺工", "缺料狀況恐延續到明年", "缺電", "能源與環保壓力", "衰退", "財務困難", "動向不明", "動盪", "國內主要仲介9月房市交易量大幅衰退", "國民黨堅決反對政府補貼台電2000億元預算", "國安危機", "國安新風險", "將改變就業市場供需", "崩盤", "推升調漲票價壓力", "推升壓力", "混亂", "產能過剩", "產業發展痛點", "通膨", "通膨壓力", "通縮", "通縮預期", "通縮壓力", "造成影響", "陷入通縮", "勞動部反對合併案", "幅度大", "復甦無望", "悲觀", "惡化", "惡性循環", "提案擱置", "最沉重的一季", "無法兌現", "無法達成共識", "短收", "窘境", "裁員", "貿易緊張", "貿易緊張升溫", "貿易緊張局勢", "貿易摩擦", "貿易摩擦", "貿易衝突", "跌幅擴大", "進口來源單一而降低了其韌性", "須持續觀察", "須留意", "匯率震盪", "意見不一", "感到不耐", "會把產業搞掉", "經濟不振", "經濟未有起色", "經濟低迷", "經濟持續低迷", "落空", "解僱", "解僱全體員工", "資本逃離", "資料短缺", "跳票", "違反美國出口管制", "違反國際貿易規則", "實務脫節", "弊案", "疑慮", "疑慮升溫", "算力有限", "綠色通膨", "緊張", "緊繃", "障礙", "需求疲軟", "需求減弱", "增加不著陸的可能性", "審慎保守", "審慎權衡", "影響房市", "影響信心", "影響員工失業", "影響產能", "憂心", "憂慮", "暴跌", "暴漲暴跌", "潛在風險", "罷工", "衝擊房市", "衝擊變數", "戰爭風險", "擋總預算", "擔憂", "擔憂升溫", "融資不順", "選前風險意識升高", "選情失利", "駭客攻擊", "壓倒性看壞", "應謹慎避險", "環境複雜嚴峻", "糟糕", "縮減產能", "縮減規模", "總統大選", "虧損加劇", "趨勢往下", "趨緊縮", "避險情緒升高", "隱憂", "謹慎情緒", "轉為緊縮", "雙方仍未達成協議", "曝險創下新高", "邊緣化", "關稅反制", "關稅翻倍", "難以扭轉", "難題", "嚴重的影響", "嚴重削弱", "嚴重衝擊", "競爭力下降", "繼續下跌", "續崩", "變化莫測", "觀望氛圍", "觀望氣氛", "陷入困境", "高額關稅", "難以預測", "需要長期觀察"]


def find_matched_keywords(text, keywords):
    return [kw for kw in keywords if kw in text]

def check_epu(text):
    matched_e = find_matched_keywords(text, E_TERMS)
    matched_p = find_matched_keywords(text, P_TERMS)
    matched_u = find_matched_keywords(text, U_TERMS)
    is_epu = bool(matched_e and matched_p and matched_u)
    return is_epu, matched_e, matched_p, matched_u

def extract_date_from_filename(fname):
    match = re.match(r"(\d{4}-\d{2}-\d{2})", fname)
    return match.group(1) if match else None

def run_check(base_dir):
    base_dir = Path(base_dir)

    # ✅ 報社合併對照表
    merged_source_map = {
        "聯合國際": "聯合報",
        "聯合產經": "聯合報",
        "自由時報_財經": "自由時報",
        "自由時報_國際": "自由時報",
    }

    # 結構：{報社: {日期: [文章條目]}}
    grouped_data = defaultdict(lambda: defaultdict(list))

    for source_dir in base_dir.iterdir():
        if not source_dir.is_dir():
            continue
        folder_name = source_dir.name
        if folder_name == "EPU匯出結果" or re.match(r"\d{4}-\d{2}-\d{2}", folder_name):
            continue

        unified_name = merged_source_map.get(folder_name, folder_name)

        for file in source_dir.glob("*.txt"):
            date_str = extract_date_from_filename(file.name)
            if not date_str:
                continue
            with open(file, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
                is_epu, matched_e, matched_p, matched_u = check_epu(text)
                grouped_data[unified_name][date_str].append({
                    "source": folder_name,
                    "filename": file.name,
                    "is_epu": is_epu,
                    "matched_e": "、".join(matched_e),
                    "matched_p": "、".join(matched_p),
                    "matched_u": "、".join(matched_u),
                    })
                

    # ✅ 每家報社每個日期獨立輸出 Excel 統計
    for paper_name, date_dict in grouped_data.items():
        for date, entries in sorted(date_dict.items()):
            rows = []
            epu_yes = 0
            for entry in entries:
                if entry["is_epu"]:
                    epu_yes += 1
                rows.append({
                "日期": date,
                "來源": entry["source"],
                "檔名": entry["filename"],
                "是否符合 EPU": "✔ 是" if entry["is_epu"] else "✘ 否",
                "E關鍵字": entry["matched_e"],
                "P關鍵字": entry["matched_p"],
                "U關鍵字": entry["matched_u"]
            })
            epu_no = len(entries) - epu_yes
            ratio = round(epu_yes / len(entries) * 100, 2) if entries else 0

            # ✅ 儲存位置：整合結果/EPU匯出結果/YYYY-MM-DD/
            output_root = base_dir / "EPU匯出結果" / date
            output_root.mkdir(parents=True, exist_ok=True)
            filename = f"{date}_{paper_name}_EPU檢查結果.xlsx"
            out_path = output_root / filename

            df = pd.DataFrame(rows)
            df.to_excel(out_path, index=False)

            # 統計摘要與圖表
            wb = load_workbook(out_path)
            ws = wb.active
            summary_row = len(rows) + 3
            ws[f"A{summary_row}"] = "統計摘要"
            ws[f"A{summary_row+1}"] = "✔ 符合 EPU"
            ws[f"B{summary_row+1}"] = epu_yes
            ws[f"A{summary_row+2}"] = "✘ 不符合 EPU"
            ws[f"B{summary_row+2}"] = epu_no
            ws[f"A{summary_row+3}"] = "EPU 比例"
            ws[f"B{summary_row+3}"] = f"{ratio}%"

            chart = BarChart()
            chart.title = "EPU 統計圖"
            chart.add_data(Reference(ws, min_col=2, min_row=summary_row+1, max_row=summary_row+2), titles_from_data=False)
            chart.set_categories(Reference(ws, min_col=1, min_row=summary_row+1, max_row=summary_row+2))
            chart.y_axis.title = "文章數"
            chart.x_axis.title = "分類"
            ws.add_chart(chart, f"D{summary_row+1}")

            wb.save(out_path)
            logging.info(f"📁 已儲存至：{out_path}")
if __name__ == "__main__":
    run_check("整合結果")
